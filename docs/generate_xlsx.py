"""Generate CardioKB documentation Excel workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
flag_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

def style_data(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

def auto_width(ws, max_width=50):
    for col in range(1, ws.max_column + 1):
        width = 0
        for row in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row, column=col).value or "")
            width = max(width, min(len(val) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max(width, 12)

# ═══════════════════════════════════════════
# Sheet 1: Data Sources
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "Data Sources"

sources_headers = ["#", "Database", "URL", "Access Type", "Have Access?", "Parser Status"]
ws1.append(sources_headers)

sources = [
    [1, "ClinicalTrials.gov", "https://clinicaltrials.gov/api/v2/studies", "Public API", "Yes", "Working (8,322 trials; skipped in --skip-download runs since it relies on a live API)"],
    [2, "ClinPGx (PharmGKB successor)", "https://api.clinpgx.org/v1/data/", "Public API", "Yes", "Working"],
    [3, "OMIM", "https://api.omim.org/api", "API key required", "No (OMIM_API_KEY)", "Working (credential-gated)"],
    [4, "DisGeNET", "https://api.disgenet.com/api/v1", "API key optional", "Partial (file fallback)", "Working"],
    [5, "NCBI Gene", "https://ftp.ncbi.nlm.nih.gov/gene/DATA/GENE_INFO/Mammalia/Homo_sapiens.gene_info.gz", "Public FTP", "Yes", "Working"],
    [6, "DoRothEA (OmniPath)", "https://omnipathdb.org/interactions?datasets=dorothea", "Public API", "Yes", "Working"],
    [7, "DrugBank", "https://go.drugbank.com/releases/", "Login required", "No (DRUGBANK_USERNAME, DRUGBANK_PASSWORD)", "Working (credential-gated)"],
    [8, "AOP-DB", "https://gaftp.epa.gov/EPADataCommons/ORD/AOP-DB/", "MySQL login required", "No (MYSQL_USERNAME, MYSQL_PASSWORD, MYSQL_DB_NAME)", "Working (credential-gated)"],
    [9, "Disease Ontology", "https://raw.githubusercontent.com/DiseaseOntology/HumanDiseaseOntology/main/src/ontology/doid.obo", "Public", "Yes", "Working"],
    [10, "Gene Ontology", "http://current.geneontology.org/ontology/go-basic.obo", "Public", "Yes", "Working"],
    [11, "Uberon", "http://purl.obolibrary.org/obo/uberon.obo", "Public", "Yes", "Working"],
    [12, "SIDER", "https://raw.githubusercontent.com/dhimmel/SIDER4/.../data/", "Public", "Yes", "Working"],
    [13, "LINCS L1000", "https://raw.githubusercontent.com/dhimmel/lincs/.../data/", "Public", "Yes", "Working"],
    [14, "PubTator Central", "https://ftp.ncbi.nlm.nih.gov/pub/lu/PubTatorCentral/", "Public FTP", "Yes", "Working"],
    [15, "CTD", "http://ctdbase.org/reports/CTD_chem_gene_ixns.tsv.gz", "Public", "Yes", "Working"],
    [16, "Bgee", "https://www.bgee.org/ftp/current/download/calls/expr_calls/", "Public FTP", "Yes", "Working"],
    [17, "Hetionet (precomputed)", "https://github.com/hetio/hetionet/raw/main/hetnet/tsv/hetionet-v1.0-edges.sif.gz", "Public", "Yes", "Working"],
    [18, "MeSH", "https://nlmpubs.nlm.nih.gov/projects/mesh/MESH_FILES/xmlmesh/desc2026.xml", "Public", "Yes", "Working (966 symptom nodes)"],
    [19, "MEDLINE Cooccurrence", "https://raw.githubusercontent.com/hetio/medline/main/data/", "Public", "Yes", "Working (3,357 DpS + 3,602 DlA + 543 DrD edges)"],
    [20, "DrugCentral", "https://unmtid-dbs.net/download/drugcentral.dump.11012023.sql.gz", "Public", "Yes", "Working (12,047 treats + 2,525 palliates)"],
    [21, "GWAS Catalog", "https://ftp.ebi.ac.uk/pub/databases/gwas/releases/latest/gwas-catalog-associations_ontology-annotated-full.zip", "Public", "Yes", "Working (760,270 gene-disease associations)"],
    [22, "BindingDB", "https://www.bindingdb.org/rwd/bind/downloads/BindingDB_All_202603_tsv.zip", "Public", "Yes", "Working (1,632,198 drug-gene bindings)"],
]
for row in sources:
    ws1.append(row)

style_header(ws1, len(sources_headers))
style_data(ws1)
auto_width(ws1)

# ═══════════════════════════════════════════
# Sheet 2: TSV File Inventory
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("TSV File Inventory")

tsv_headers = ["#", "Source", "Filename", "Rows", "Columns", "Notes"]
ws2.append(tsv_headers)

tsv_files = [
    [1, "clinicaltrials", "clinical_trials.tsv", 8313, "trial_id, title, intervention_name, condition, phase, status, source_database"],
    [2, "clinicaltrials", "trial_studies_condition.tsv", 19004, "trial_id, condition"],
    [3, "clinicaltrials", "trial_tests_intervention.tsv", 12512, "trial_id, intervention_name"],
    [4, "clinpgx", "clinical_annotations.tsv", 454, "annotation_id, gene, drug, variant, evidence_level, types, source_database"],
    [5, "clinpgx", "drug_labels.tsv", 378, "label_id, name, drug, gene, source, biomarker_status, testing, alternate_drug_available, source_database"],
    [6, "clinpgx", "gene_info.tsv", 17, "clinpgx_id, gene_symbol, gene_name, chromosome, cpic_gene, vip_tier, source_database"],
    [7, "clinpgx", "variants.tsv", 1060, "variant_id, variant_name, gene, chromosome, position, change_classification, source_database"],
    [8, "clinpgx", "variant_in_gene.tsv", 1060, "variant_id, gene"],
    [9, "ncbigene", "genes.tsv", 193790, "tax_id, GeneID, Symbol, LocusTag, Synonyms, dbXrefs, chromosome, map_location, description, type_of_gene, Symbol_from_nomenclature_authority, Full_name_from_nomenclature_authority, Nomenclature_status, Other_designations, Modification_date, Feature_type, xref_MIM, xref_HGNC, xref_Ensembl, source_database", "LocusTag is blank ('-') for all human genes — only used for bacterial/archaeal genomes. Feature_type is blank for most rows — only populated for non-coding/regulatory genes (enhancers, silencers, etc.). Both are expected from the NCBI source file."],
    [10, "dorothea", "transcription_factors.tsv", 367, "tf_symbol, node_type, source_database"],
    [11, "dorothea", "tf_gene_interactions.tsv", 15267, "tf_symbol, target_gene, tf_uniprot, target_uniprot, confidence, curation_effort, mode_of_regulation, mor_score, is_directed, relationship, source_database"],
    [12, "disease_ontology", "disease_nodes.tsv", 12012, "doid, name, definition, synonyms"],
    [13, "disease_ontology", "disease_xrefs.tsv", 38573, "doid, xref"],
    [14, "gene_ontology", "biological_process_nodes.tsv", 24547, "go_id, name, definition, namespace"],
    [15, "gene_ontology", "molecular_function_nodes.tsv", 10123, "go_id, name, definition, namespace"],
    [16, "gene_ontology", "cellular_component_nodes.tsv", 4069, "go_id, name, definition, namespace"],
    [17, "gene_ontology", "gene_bp_associations.tsv", 158164, "gene_symbol, go_id, evidence, relationship"],
    [18, "gene_ontology", "gene_mf_associations.tsv", 106409, "gene_symbol, go_id, evidence, relationship"],
    [19, "gene_ontology", "gene_cc_associations.tsv", 111869, "gene_symbol, go_id, evidence, relationship"],
    [20, "uberon", "anatomy_nodes.tsv", 15633, "uberon_id, name, definition"],
    [21, "sider", "side_effect_nodes.tsv", 5734, "umls_id, name, source"],
    [22, "sider", "compound_causes_side_effect.tsv", 138944, "drugbank_id, umls_id, name, source, relationship"],
    [23, "lincs", "compound_upregulates_gene.tsv", 33024, "drugbank_id, entrez_gene_id, z_score, method, unbiased, source, sourceDatabase"],
    [24, "lincs", "compound_downregulates_gene.tsv", 36163, "drugbank_id, entrez_gene_id, z_score, method, unbiased, source, sourceDatabase"],
    [25, "lincs", "gene_regulates_gene.tsv", 267812, "source_gene, target_gene, z_score, subtypes, method, unbiased, source, sourceDatabase"],
    [26, "pubtator", "disease_disease_cooccurrence.tsv", 9367474, "disease1_id, disease2_id, pmid_count, relationship, source"],
    [27, "pubtator", "gene_disease_literature.tsv", 59879417, "gene_id, disease_id, pmid_count, relationship, source"],
    [28, "ctd", "chemical_increases_expression.tsv", 346081, "chemical_name, chemical_id, gene_symbol, gene_id, interaction_type, pubmed_ids, source, relationship"],
    [29, "ctd", "chemical_decreases_expression.tsv", 330934, "chemical_name, chemical_id, gene_symbol, gene_id, interaction_type, pubmed_ids, source, relationship"],
    [30, "bgee", "bodypart_overexpresses_gene.tsv", 3200, "gene_id, gene_name, anatomy_id, anatomy_name, expression_score, expression_rank, source, relationship"],
    [31, "bgee", "bodypart_underexpresses_gene.tsv", 6605912, "gene_id, gene_name, anatomy_id, anatomy_name, expression_score, expression_rank, source, relationship"],
    [32, "hetionet_precomputed", "gene_covaries.tsv", 61690, "gene1_symbol, gene2_symbol, relationship, source"],
    [33, "hetionet_precomputed", "gene_regulates.tsv", 265672, "gene1_symbol, gene2_symbol, relationship, source"],
    [34, "hetionet_precomputed", "gene_interacts.tsv", 147164, "gene1_symbol, gene2_symbol, relationship, source"],
    [35, "hetionet_precomputed", "drug_causes_effect.tsv", 138944, "drug_id, effect_id, relationship, source"],
    [36, "mesh", "symptom_nodes.tsv", 966, "mesh_id, name, definition, tree_numbers"],
    [37, "medline", "disease_symptom_cooccurrence.tsv", 3357, "doid_code, mesh_id, p_fisher, cooccurrence, enrichment, source, unbiased, license, sourceDatabase"],
    [38, "medline", "disease_anatomy_cooccurrence.tsv", 3602, "doid_code, uberon_id, p_fisher, cooccurrence, enrichment, source, unbiased, license, sourceDatabase"],
    [39, "medline", "disease_disease_cooccurrence.tsv", 543, "doid_code_0, doid_code_1, p_fisher, cooccurrence, enrichment, source, unbiased, license, sourceDatabase"],
    [40, "drugcentral", "drug_treats_disease.tsv", 12047, "struct_id, concept_id, relationship_name, concept_name, umls_cui, snomed_full_name, source"],
    [41, "drugcentral", "drug_palliates_disease.tsv", 2525, "struct_id, concept_id, relationship_name, concept_name, umls_cui, snomed_full_name, source"],
    [42, "gwas", "gene_disease_gwas.tsv", 760270, "gene_symbol, disease_trait, mapped_trait, mapped_trait_uri, p_value, study_accession, pubmed_id, relationship, source"],
    [43, "bindingdb", "drug_binds_gene.tsv", 1632198, "ligand_name, drugbank_id, pubchem_cid, target_name, uniprot_id, affinity_nm, affinity_type, relationship, source"],
]
for row in tsv_files:
    ws2.append(row)

# Format row counts with comma separators
for r in range(2, ws2.max_row + 1):
    cell = ws2.cell(row=r, column=4)
    cell.number_format = "#,##0"

style_header(ws2, len(tsv_headers))
style_data(ws2)
auto_width(ws2)

# ═══════════════════════════════════════════
# Sheet 3: Flags / Issues
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Flags")

flag_headers = ["#", "File", "Issue", "Severity"]
ws3.append(flag_headers)

flags = [
    [1, "clinpgx/gene_info.tsv", "Only 17 rows — verify all CVD pharmacogenes were captured", "Low"],
    [2, "clinpgx/variants.tsv", "chromosome and position columns are empty in sampled rows", "Medium"],
    [3, "clinpgx/drug_labels.tsv", "testing column contains raw Python dict strings instead of parsed values", "Medium"],
    [4, "bgee/bodypart_overexpresses_gene.tsv", "Only 3,200 rows vs 6.6M underexpressed (2000x ratio) — possible filter/threshold issue", "Medium"],
    [5, "hetionet_precomputed/gene_covaries.tsv", "Columns named gene1_symbol/gene2_symbol but contain Entrez IDs, not symbols", "Medium"],
    [6, "hetionet_precomputed/gene_regulates.tsv", "Same column name vs Entrez ID mismatch", "Medium"],
    [7, "hetionet_precomputed/gene_interacts.tsv", "Same column name vs Entrez ID mismatch", "Medium"],
    [8, "clinicaltrials/clinical_trials.tsv", "Skipped in --skip-download runs (live API); works fine on normal runs (8,322 trials confirmed)", "Low"],
]
for row in flags:
    ws3.append(row)

# Highlight severity
severity_colors = {
    "Medium": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Low": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
}
for r in range(2, ws3.max_row + 1):
    sev = ws3.cell(row=r, column=4).value
    if sev in severity_colors:
        for c in range(1, ws3.max_column + 1):
            ws3.cell(row=r, column=c).fill = severity_colors[sev]

style_header(ws3, len(flag_headers))
style_data(ws3)
auto_width(ws3)

# ── Save ──
out = "/Users/nawaza/Desktop/Cardio-KB/docs/cardiokb_data_inventory.xlsx"
wb.save(out)
print(f"Saved to {out}")
